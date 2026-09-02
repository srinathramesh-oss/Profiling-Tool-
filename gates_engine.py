#!/usr/bin/env python3
"""
LODHA GCR — DETERMINISTIC ENGINE, OAuth edition (no service account, no
Apps Script bridge, no Google Cloud IAM)

Authenticates as YOU, via the refresh token in token.json, produced once
through the OAuth consent flow. Reads and writes the Sheet and Drive with
your own existing access — nothing is shared with any other identity, so
the Workspace admin's external-sharing policy never enters into it, and
there is no Apps Script deployment-access wall to fight.

This file does no research and never calls a model. Research is done by
the routine's own Claude Code agent (see ROUTINE_PROMPT.md), which is what
keeps model usage on your subscription quota rather than a separate API
bill. This file only does the parts that must be exact code: sheet reads
and writes, the gate math, the evidence workbook, the Drive upload.

CLI (unchanged contract from earlier versions — same subcommands
ROUTINE_PROMPT.md already calls):

  read_queue
  claim --row N
  evaluate --sub '<json>' --anchor '<json>' --ledger '<json>'
  workbook --sub '<json>' --anchor '<json>' --ledger '<json>' --notes '<json>'
  writeback --row N --verdict V --confidence C --summary S --reasoning R
             --gaps '["...","..."]' --escalations E --evidence_url U
  fail --row N --reason "..."

ENVIRONMENT
  GCR_SHEET_ID       required — the spreadsheet ID from its URL
  GCR_DRIVE_FOLDER   optional — Drive folder ID for evidence workbooks
  GOOGLE_OAUTH_TOKEN the full contents of token.json, as one line

  pip install google-auth google-auth-oauthlib google-api-python-client openpyxl
"""

import argparse, json, os, re, sys
from datetime import datetime, timezone

from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request as GoogleRequest
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SHEET_ID  = os.environ.get("GCR_SHEET_ID", "")
FOLDER_ID = os.environ.get("GCR_DRIVE_FOLDER", "")
LEADS     = "Leads"
STALE_MINUTES = 30

ST = dict(QUEUED="QUEUED", IDENTITY_CHECK="IDENTITY_CHECK", CONFIRMED="CONFIRMED",
          ENRICHING="ENRICHING", PENDING="PENDING_APPROVAL", FAILED="FAILED")

CONFIG_DEFAULTS = dict(identityMin=0.60, baseTicketCr=200, baseTurnoverFloorCr=500,
                       baseNetWorthFloorCr=200, capacityRedBelowRatio=0.5,
                       developerCarveOutCr=500, materialityCr=1)
# (label, ticket Cr, matching floor-area band) — area at Rs 20,000/sq ft
BANDS = [("\u20B9200 \u2013 300 Cr", 200, "1,00,000 \u2013 1,50,000 sq ft"),
         ("\u20B9300 \u2013 500 Cr", 300, "1,50,000 \u2013 2,50,000 sq ft"),
         ("\u20B9500 \u2013 750 Cr", 500, "2,50,000 \u2013 3,75,000 sq ft"),
         ("Above \u20B9750 Cr",       750, "Above 3,75,000 sq ft")]

TAX_MATTER = re.compile(r"\b(gst|goods and services tax|income[- ]tax|tax demand|tax notice|show[- ]cause|assessment order|adjudication order|input tax credit|customs|excise|vat|service tax|tds|transfer pricing)\b", re.I)
CONVICTION = re.compile(r"\b(convicted|conviction|found guilty|pleaded guilty|sentenced|rigorous imprisonment|criminal breach of trust|debarred|disqualified as a director|barred from (?:trading|the securities market|holding))\b", re.I)
EXONERATION = re.compile(r"\b(acquitted|acquittal|exonerated|discharged|quashed|set aside|closure report|no case (?:was )?made out)\b", re.I)
CONSUMER = re.compile(r"\b(consumer (?:court|forum|commission|dispute|redressal)|district commission|state commission|ncdrc|deficiency in service|small claims)\b", re.I)

def _txt(f): return f"{(f or {}).get('value','')} {(f or {}).get('note','')}"
def is_tax(f): return bool(f) and bool(TAX_MATTER.search(_txt(f)))
def is_conviction(f): return bool(f) and bool(CONVICTION.search(_txt(f)))
def is_exonerated(f): return bool(f) and bool(EXONERATION.search(_txt(f)))
def is_weak_source(f): return bool(f) and str(f.get("tier","")).lower() == "unverified"
def is_foreign_affiliate(f):
    if not f or str(f.get("jurisdiction","")).lower() != "foreign": return False
    return str(f.get("about","subject")).lower() != "subject"
def amount_cr(text):
    t = str(text or "").replace(",", "")
    m = re.search(r"(?:rs\.?|inr|\u20B9)\s*([\d.]+)\s*(lakh\s*crore|crore|cr\b|lakhs?)?", t, re.I)
    if not m: return None
    try: n = float(m.group(1))
    except ValueError: return None
    unit = (m.group(2) or "").lower()
    if unit.startswith("lakh crore"): return n * 100000
    if unit.startswith("crore") or unit == "cr": return n
    if unit.startswith("lakh"): return n / 100
    return n / 10000000
def is_minor(f, cfg):
    if not f: return False
    if CONSUMER.search(_txt(f)): return True
    a = amount_cr(_txt(f))
    return a is not None and a < cfg["materialityCr"]

def fmt_cr(cr):
    if cr is None: return ""
    n = round(cr) if cr >= 100 else round(cr, 1)
    s = str(n); ip, _, dec = s.partition(".")
    if len(ip) > 3:
        last3, rest = ip[-3:], ip[:-3]
        rest = re.sub(r"\B(?=(\d{2})+(?!\d))", ",", rest)
        ip = rest + "," + last3
    return "\u20B9" + ip + ("." + dec if dec else "") + " Cr"
def ticket_for(sub):
    area = (sub or {}).get("area", "")
    for lbl, tk, ar in BANDS:
        if ar == area: return tk
    budget = (sub or {}).get("budget", "")
    return next((b[1] for b in BANDS if b[0] == budget), BANDS[0][1])
NF, CPSRC = "Not found in public sources", "Channel partner form"
MONS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

def now_iso(): return datetime.now(timezone.utc).isoformat()
def dmy(dt=None):
    dt = dt or datetime.now()
    return f"{dt.day:02d}/{MONS[dt.month-1]}/{dt.year}"
def col(idx0): return get_column_letter(idx0 + 1)

# ---------------------------------------------------------------- auth

def gauth():
    """Loads token.json's contents from GOOGLE_OAUTH_TOKEN, refreshes the
    access token if needed, and returns Sheets + Drive clients. No service
    account file, no IAM, no Drive/Sheets sharing with anyone — this
    credential is you."""
    raw = os.environ.get("GOOGLE_OAUTH_TOKEN", "").strip()
    if not raw:
        print("GOOGLE_OAUTH_TOKEN is not set (paste the contents of token.json as one line)", file=sys.stderr)
        sys.exit(2)
    data = json.loads(raw)
    scopes = data.get("scopes") or ["https://www.googleapis.com/auth/spreadsheets",
                                    "https://www.googleapis.com/auth/drive"]
    creds = Credentials(
        token=data.get("token"),
        refresh_token=data["refresh_token"],
        token_uri=data.get("token_uri", "https://oauth2.googleapis.com/token"),
        client_id=data["client_id"],
        client_secret=data["client_secret"],
        scopes=scopes,
    )
    # The access token in token.json (if any) is almost certainly expired by
    # the time a scheduled run fires; refresh unconditionally rather than
    # trusting a timestamp that may not even be present.
    creds.refresh(GoogleRequest())
    return build("sheets", "v4", credentials=creds), build("drive", "v3", credentials=creds)

def read_all(sheets, tab):
    return sheets.spreadsheets().values().get(spreadsheetId=SHEET_ID, range=tab).execute().get("values", [])

def write_cells(sheets, updates):
    body = {"valueInputOption": "RAW", "data": [{"range": r, "values": [[v]]} for r, v in updates]}
    sheets.spreadsheets().values().batchUpdate(spreadsheetId=SHEET_ID, body=body).execute()

def load_anchors(sheets):
    out = []
    try:
        for row in read_all(sheets, "Anchors")[1:]:
            if len(row) >= 2 and row[0].strip() and row[1].strip():
                out.append(dict(occupier=row[0].strip(), line=row[1].strip()))
    except Exception:
        pass
    return out

def load_config(sheets):
    cfg = dict(CONFIG_DEFAULTS)
    try:
        for row in read_all(sheets, "Config")[1:]:
            if len(row) >= 2 and row[0].strip() in cfg:
                try: cfg[row[0].strip()] = float(row[1])
                except ValueError: pass
    except Exception:
        pass
    return cfg

def leads_header(sheets):
    data = read_all(sheets, LEADS)
    return data[0], data[1:], {h: i for i, h in enumerate(data[0])}

# ---------------------------------------------------------------- gates (verbatim port of Code.gs)

def cr_from(text):
    if not text: return None
    t = str(text).replace(",", "")
    m = re.search(r"([\d.]+)\s*lakh\s*crore", t, re.I)
    if m: return float(m.group(1)) * 100000
    m = re.search(r"([\d.]+)\s*(crore|cr\b)", t, re.I)
    return float(m.group(1)) if m else None

def evaluate_gates(ledger, anchor, budget, cfg, sub=None):
    G = []
    def get(fid):
        f = ledger.get(fid)
        return f if f and f.get("status") == "found" else None
    def add(fid, label, result, detail, escalate=False):
        G.append(dict(id=fid, label=label, result=result, detail=detail, escalate=escalate))

    conf = float((anchor or {}).get("confidence") or 0)
    add("identity", "Identity resolved", "pass" if conf >= cfg["identityMin"] else "unresolved",
        f"confidence {conf}" if conf else "identity step returned nothing")

    def required_for(ticket):
        k = ticket / cfg["baseTicketCr"]
        return cfg["baseTurnoverFloorCr"] * k, cfg["baseNetWorthFloorCr"] * k

    turn = get("6b");  cr  = cr_from(turn["value"]) if turn else None
    worth = get("3a"); wcr = cr_from(worth["value"]) if worth else None
    liq = get("d2");   liq_cr = cr_from(liq["value"]) if liq else None
    if liq_cr is not None and (wcr is None or liq_cr > wcr): wcr = liq_cr
    elif liq and wcr is None: wcr = cfg["baseNetWorthFloorCr"]

    ticket = ticket_for(sub or {"budget": budget})
    need_t, need_w = required_for(ticket)
    best = None
    for lbl, tk, _ar in BANDS:
        nt, nw = required_for(tk)
        if ((cr is not None and cr >= nt) or (wcr is not None and wcr >= nw)) and (best is None or tk > best[1]):
            best = (lbl, tk)
    routes = []
    if cr is not None and cr >= need_t: routes.append(f"turnover {fmt_cr(cr)}")
    if worth and wcr is not None and wcr >= need_w: routes.append(f"net worth {fmt_cr(wcr)}")
    if liq and wcr is not None and wcr >= need_w and not worth: routes.append("liquidity event")
    need_txt = f"needs {fmt_cr(need_t)}"

    if routes:
        add("capacity", "Capacity against ticket", "pass", ", ".join(routes) + " \u00b7 " + need_txt)
    elif best:
        add("capacity", "Capacity against ticket", "qualifies lower",
            f"supports {best[0]}, not the {fmt_cr(ticket)} asked for", True)
    elif cr is None and wcr is None:
        add("capacity", "Capacity against ticket", "not established", "no turnover, net worth or liquidity event found")
    elif cr is not None and cr < cfg["baseTurnoverFloorCr"] * cfg["capacityRedBelowRatio"]:
        add("capacity", "Capacity against ticket", "far below",
            f"{fmt_cr(cr)} turnover \u00b7 smallest unit needs {fmt_cr(cfg['baseTurnoverFloorCr'])}")
    else:
        add("capacity", "Capacity against ticket", "short", f"{fmt_cr(cr if cr is not None else wcr)} \u00b7 {need_txt}")

    for fid, label in [("g1","Politically exposed person"),("g2","Real estate developer interest"),
                       ("g3","Practising lawyer"),("g4","Journalist"),("g5","Public visibility")]:
        f = get(fid)
        if not f: add(fid, label, "nothing found", "no match in the sources searched"); continue
        if fid == "g2" and cr is not None and cr >= cfg["developerCarveOutCr"]:
            add(fid, label, "exception applies", f"{f['value']} \u2014 above the {fmt_cr(cfg['developerCarveOutCr'])} carve-out"); continue
        add(fid, label, "for the screening authority", f["value"], True)

    # tax and regulatory: to the committee, never a Red on its own
    f = get("e4")
    if not f: add("e4", "Regulatory or tax proceedings", "nothing found", "no match in the sources searched")
    else:
        sev = (f.get("severity") or "mention").lower()
        if sev == "mention": add("e4", "Regulatory or tax proceedings", "noted", f"{f['value']} \u2014 press mention only")
        else: add("e4", "Regulatory or tax proceedings", "for the committee",
                  f"{f['value']} \u2014 " + ("concluded" if sev == "finding" else "under challenge or unresolved"), True)

    for fid, label in [("5c","Litigation, individual or co-directors"),
                       ("e5","Default, NPA or insolvency"),("5d","Litigation, family members")]:
        f = get(fid)
        if not f: add(fid, label, "nothing found", "no match in the sources searched"); continue
        sev = (f.get("severity") or "mention").lower(); about = (f.get("about") or "subject").lower()
        own = about in ("subject", "company")
        if is_conviction(f) and not is_tax(f): sev = "finding"
        if sev == "finding" and is_weak_source(f): sev = "allegation"
        if is_foreign_affiliate(f):
            add(fid, label, "noted", f"{f['value']} \u2014 a foreign affiliate, not the Indian entity or the buyer"); continue
        if is_exonerated(f) and not is_conviction(f): sev = "mention"
        elif is_exonerated(f) and is_conviction(f): sev = "allegation"
        if is_minor(f, cfg) and not is_conviction(f):
            add(fid, label, "noted", f"{f['value']} \u2014 too small to bear on a purchase at this level"); continue
        if is_tax(f):
            if sev == "mention": add(fid, label, "noted", f"{f['value']} \u2014 a tax matter, press mention only")
            else: add(fid, label, "for the committee", f"{f['value']} \u2014 a tax matter, for the committee to weigh", True)
        elif sev == "finding" and own: add(fid, label, "disqualifying", f"{f['value']} \u2014 concluded, against the buyer")
        elif sev == "finding": add(fid, label, "for the committee", f"{f['value']} \u2014 concluded, but against a {about}", True)
        elif sev == "allegation": add(fid, label, "for the committee", f"{f['value']} \u2014 unproven, {about}", True)
        else: add(fid, label, "noted", f"{f['value']} \u2014 press mention only, {about}")

    anchors = cfg.get("anchors") or []
    if anchors:
        ind = str((sub or {}).get("industry") or (anchor or {}).get("industry") or "").lower()
        f6g = get("6g")
        if not ind and f6g: ind = str(f6g["value"]).lower()
        hit = next((a for a in anchors if ind and (a["line"].lower() in ind or ind in a["line"].lower())), None)
        if hit: add("anchorConflict", "Conflict with an existing occupier", "for the committee",
                    f"same line of business as {hit['occupier']} ({hit['line']})", True)
        elif ind: add("anchorConflict", "Conflict with an existing occupier", "nothing found", "no overlap with any anchor occupier")
        else: add("anchorConflict", "Conflict with an existing occupier", "not established", "client\u2019s line of business not established")
    return G

# ---------------------------------------------------------------- breakdown (mirrors Code.gs)

RAG_RANK = dict(grey=0, green=1, amber=2, red=3)
def state_of(result):
    if result in ("disqualifying", "far below"): return "red"
    if result in ("nothing found", "pass", "exception applies", "noted"): return "green"
    if result in ("not established", "untestable"): return "grey"
    return "amber"
def worst(states): return max(states, key=lambda s: RAG_RANK[s]) if states else "grey"
def brief(text, words=10):
    t = str(text or "").strip()
    t = re.split(r"(?<=[.;])\s", t)[0]
    t = re.split(r"\s+\u2014\s+", t)[0]
    w = t.split()
    if len(w) > words: t = " ".join(w[:words]) + "\u2026"
    return re.sub(r"[,;:\s]+$", "", t)
def breakdown(gates, ledger, anchor):
    def g(fid): return next((x for x in gates if x["id"] == fid), None)
    def get(fid):
        f = ledger.get(fid); return f if f and f.get("status") == "found" else None
    def src_of(ids):
        for i in ids:
            f = get(i)
            if f and f.get("source"): return str(f["source"])
        return ""
    anchor_src = str(((anchor or {}).get("sources") or [""])[0] or "")
    def row(label, gate, ids=()):
        return dict(label=label, state=state_of(gate["result"]), note=gate["detail"], src=src_of(ids)) if gate else None
    dims = []
    def push(name, subs):
        subs = [s for s in subs if s]
        dims.append(dict(name=name, subs=[dict(label=s["label"], state=s["state"], note=str(s.get("note") or ""), src=s.get("src") or "") for s in subs],
                         state=worst([s["state"] for s in subs])))
    idg = g("identity") or dict(result="untestable", detail="not checked")
    edu_f = get("i7")
    edu = (edu_f or {}).get("value") or (anchor or {}).get("education") or ""
    cap = g("capacity") or dict(result="untestable", detail="not checked")
    turn = get("6b"); nw = get("3a") or get("d2")
    push("Business Reputation", [
        row("Cases \u2014 client and co-directors", g("5c"), ["5c"]), row("Tax and duty disputes", g("e4"), ["e4"]),
        row("Developer connection", g("g2"), ["g2"]), row("Conflict with an existing occupier", g("anchorConflict"), ["6g"])])
    push("Personal Image", [
        dict(label="Identity confirmed", state=state_of(idg["result"]), note=idg["detail"], src=anchor_src),
        dict(label="Education", state="green" if edu else "grey", note=edu or "not established",
             src=str(edu_f.get("source") or "") if edu_f else (anchor_src if edu else "")),
        row("Political exposure", g("g1"), ["g1"]), row("Practising lawyer", g("g3"), ["g3"]),
        row("Journalist", g("g4"), ["g4"]), row("Public attention", g("g5"), ["g5"]),
        row("Cases \u2014 immediate family", g("5d"), ["5d"])])
    push("Financial Capacity", [
        dict(label="Turnover", state="green" if turn else "grey", note=turn["value"] if turn else "not established", src=str(turn.get("source") or "") if turn else ""),
        dict(label="Net worth or liquidity", state="green" if nw else "grey", note=nw["value"] if nw else "not established", src=str(nw.get("source") or "") if nw else ""),
        dict(label="Fit to the ticket asked for", state=state_of(cap["result"]), note=cap["detail"], src=src_of(["6b","3a","d2"])),
        row("Default or insolvency", g("e5"), ["e5"])])
    return dims
def trim_gaps(gaps):
    out = []
    for gtext in (gaps or [])[:4]:
        t = str(gtext or "").strip()
        t = re.split(r"(?<=[.;])\s", t)[0]
        t = re.split(r"\s+\u2014\s+|,\s+(?:but|since|though|although|and the committee)\s", t)[0]
        t = re.sub(r"[.,;:\s]+$", "", t)
        if t: out.append(t)
    return out

def rate(G):
    by = {g["id"]: g for g in G}
    if any(g["result"] == "disqualifying" for g in G): return "Red"
    if by["capacity"]["result"] == "far below": return "Red"
    if any(g["escalate"] for g in G): return "Amber"
    if by["identity"]["result"] != "pass": return "Amber"
    if by["capacity"]["result"] in ("qualifies lower", "short"): return "Amber"
    return "Green"

# ---------------------------------------------------------------- evidence xlsx (Alibaug format)

TITLE_FILL, BAND_FILL, REMARK_FILL = (PatternFill("solid", fgColor=c) for c in ("FABF8F","FCD5B4","FBF8F2"))
INK  = Font(name="Calibri", size=11, bold=True, color="1F4E79")
BOLD = Font(name="Calibri", size=11, bold=True)
BASE = Font(name="Calibri", size=10)
MUTEF = Font(name="Calibri", size=10, color="808080")
META = Font(name="Calibri", size=9, italic=True, color="808080")
WRAP = Alignment(vertical="top", wrap_text=True)

def build_rows(sub, anchor, ledger, notes):
    R = []
    v = lambda x: str(x).strip() if x and str(x).strip() else ""
    def get(fid):
        f = ledger.get(fid)
        return f if f and f.get("status") == "found" else None
    a_src = ((anchor or {}).get("sources") or [""])[0]
    def line(param, value, comment="", source="", guide="", stage="In meeting"):
        val = v(value); found = bool(val)
        R.append(dict(t="line", found=found, c=[param, val if found else NF, comment, (source if found else ""), guide, ("" if found else stage)]))
    TIER = dict(primary="filing or order", reported="press report", unverified="unverified source")
    def src_of(f):
        if not f or not f.get("value"): return ""
        t = TIER.get(str(f.get("tier","")).lower())
        return "  \u00b7  ".join(x for x in [f.get("source"), t] if x)
    def fld(param, fid, guide="", stage="In meeting", comment=""):
        f = get(fid)
        line(param, f["value"] if f else "", (f.get("note") if f else comment) or comment, src_of(f), guide, stage)
    def either(param, typed, anchored, guide="", comment=""):
        val = v(typed) or v(anchored or "")
        line(param, val, comment, CPSRC if v(typed) else (a_src if val else ""), guide)
    def band(t): R.append(dict(t="band", c=[t, "", "Comments", "Source of Information", "Guideline for assessing each parameter", "Filling Stage"]))
    def remark(txt, frm): R.append(dict(t="remark", c=["Remarks", v(txt), "", "", f"Open text — {frm}" if frm else "Open text — for the assessing committee", ""]))

    R.append(dict(t="title", c=["LODHA GOLF COURSE ROAD — CLIENT PROFILE"]))
    R.append(dict(t="meta", c=["Reference", sub["reference"]]))
    R.append(dict(t="meta", c=["Prepared", dmy() + " " + datetime.now().strftime("%H:%M")]))
    R.append(dict(t="meta", c=["Identity confidence", str(anchor.get("confidence")) if anchor and anchor.get("confidence") is not None else "not resolved"]))
    R.append(dict(t="meta", c=["Status", "Evidence file — not assessed"]))
    R.append(dict(t="spacer", c=[]))

    band("Client Details")
    line("Client Name", sub["clientName"], "", CPSRC)
    line("Existing Lodha Buyer?", sub.get("existing"), "Needs the CC / HPM record to confirm", CPSRC)
    line("Occupation", sub.get("role"), "", CPSRC)
    either("Company Name", sub.get("company"), anchor.get("company") if anchor else "")
    either("Office Address", sub.get("officeCity"), anchor.get("office_address") if anchor else "")
    either("Designation", sub.get("role"), anchor.get("designation") if anchor else "")
    either("Official company website link:", sub.get("website"), anchor.get("website") if anchor else "")
    remark(anchor.get("note") if anchor else "", "Identity step")

    band("Client Requirement & Source Details")
    line("Budget", sub.get("budget"), "Units start at ₹200 Cr", CPSRC)
    line("Configuration interested in", " · ".join(x for x in [v(sub.get("area")), v(sub.get("purpose"))] if x), "", CPSRC)
    line("Source", sub.get("leadSource") or "Channel Partner", "", "Intake form")
    line("Source Details ( CP Company name)",
         " · ".join(x for x in [v(sub.get("cpFirm")), v(sub.get("cpSpoc")), f"Reap {sub['reapId']}" if v(sub.get("reapId")) else ""] if x),
         "", CPSRC, "", "Before walkin")
    line("Sourcing SM", " · ".join(x for x in [v(sub.get("smName")), v(sub.get("smEmail"))] if x), "", CPSRC)
    remark(sub.get("notes"), "Channel partner")

    band("1. Primary Residence")
    fld("1a. Residential Building", "1a", "4/5 BHK")
    fld("1b. Location", "1b")
    line("1b. Owned / Rented", "", "", "", "Owned")
    fld("1c. Home market value", "1c", "20 cr+ home value")
    remark(notes.get("lifestyle"), "Residence and affiliations")

    band("2. Secondary Residences")
    fld("2a. Secondary Real Estate Investments/Locations", "2a")
    line("2b. Existing Properties in Gurugram / NCR", "")
    remark("", "")

    band("3. Net worth")
    fld("3a. Estimated Net worth:\n- Company ownership value + Other assets", "3a", "100 Cr+")
    fld("3b. Annual Income", "3b", "5 Cr+")
    fld("d2. Liquidity events", "d2", "Sale, acquisition, listing, funding round or large dividend — value and year")
    remark(notes.get("corporate"), "Company filings")

    band("4. Lifestyle")
    fld("4a. Luxury ecosystem affiliations:\n- Memberships in Golf club, Yacht club, business clubs", "4a")
    line("4b. Sponsor/Organizer/Patron in Art fairs, galleries, events", "")
    line("4c. High-profile events attended", "")
    fld("4d. Involvement in philanthropic activities or trusteeship", "4d")
    remark("", "")

    band("5. Reputation")
    R.append(dict(t="sub", c=["Industry Reputation"]))
    f5a = get("5a")
    line("5a. Seniority in the organization (MD/CEO, MD/CEO-1, MD/CEO-2)",
         f5a["value"] if f5a else (sub.get("role") or (anchor.get("designation") if anchor else "")),
         "", (f5a.get("source") if f5a else (CPSRC if sub.get("role") else a_src)), "MD/CEO minus 2 or higher")
    fld("5b. Positive reputation in the industry", "5b")
    fld("5c. Any litigations or negative news about self or other directors", "5c")
    R.append(dict(t="sub", c=["Social Reputation"]))
    fld("5d. Any litigations or negative news about family members", "5d")
    line("5e. Closed network reference check", "", "Needs the CC / HPM feed — internal lookup", "", "Delayed CAM, conduct in existing society", "Customer Care")
    hits = [get(g) for g in ("g1","g2","g3","g4","g5") if get(g)]
    line("5f. Sensitive Profile check", "; ".join(h["value"] for h in hits),
         "Community screen deliberately excluded — handled offline by the committee",
         hits[0]["source"] if hits else "", "Per policy 05.10.24 v3")
    remark("  —  ".join(x for x in [notes.get("news"), notes.get("litigation"), notes.get("sensitive")] if x), "Press, court and screening lanes")

    band("6. Entity & Occupier — Golf Course Road")
    fld("6a. Entity type", "6a")
    fld("6b. Annual turnover", "6b", "Scaled to the unit asked about")
    fld("6c. PAT trend, three years", "6c")
    fld("6d. Group / related entities", "6d")
    either("6e. Current office, owned or leased", sub.get("currentOffice"), (get("6e") or {}).get("value"))
    fld("6f. Headcount", "6f")
    either("6g. Industry / business line", sub.get("industry"), anchor.get("industry") if anchor else "")
    line("6h. Anchor-occupant business conflict", "", "Not testable — anchor list not supplied", "", "Same business line as any occupant holding 20%+", "Sales lead")
    remark("", "")

    band("7. Standing & Influence — Golf Course Road")
    fld("7a. Board or advisory positions at other companies", "i1")
    fld("7b. Industry body roles", "i2")
    fld("7c. Trusteeship or institutional association", "i3")
    fld("7d. Corporate philanthropy", "i4")
    fld("7e. Public platform", "i5")
    fld("7f. Business lineage", "i6")
    f7 = get("i7")
    edu_val = (f7 or {}).get("value") or (anchor or {}).get("education") or ""
    edu_src = src_of(f7) if f7 else (((anchor or {}).get("sources") or [""])[0])
    line("7g. Education", edu_val, "", edu_src if edu_val else "", "Degrees and institutions, where a public source states them")
    remark(notes.get("standing"), "Standing and influence")

    R.append(dict(t="spacer", c=[]))
    R.append(dict(t="note", c=["Blank cells mean nothing was found in the sources searched. That is not the same as a clear finding."]))
    return R

def write_workbook(rows, sub, out_dir="/tmp"):
    wb = Workbook(); ws = wb.active; ws.title = "Evidence"
    for i, w in enumerate([55, 31.8, 28.8, 27.4, 54.1, 18]):
        ws.column_dimensions[col(i)].width = w
    ws.freeze_panes = "D1"
    for n, r in enumerate(rows, start=1):
        c = (r["c"] + [""] * 6)[:6]
        for j, val in enumerate(c, start=1):
            cell = ws.cell(row=n, column=j, value=val or None); cell.font = BASE; cell.alignment = WRAP
        t = r["t"]
        if t == "title":
            ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=3)
            for j in range(1, 7): ws.cell(row=n, column=j).fill = TITLE_FILL
            ws.cell(row=n, column=1).font = INK
        elif t == "meta":
            ws.cell(row=n, column=1).font = META; ws.cell(row=n, column=2).font = META
        elif t == "band":
            ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=2)
            for j in range(1, 7): ws.cell(row=n, column=j).fill = BAND_FILL
            ws.cell(row=n, column=1).font = BOLD
        elif t == "sub":
            ws.cell(row=n, column=1).font = BOLD
        elif t == "note":
            ws.merge_cells(start_row=n, start_column=1, end_row=n, end_column=6)
            ws.cell(row=n, column=1).font = META
        elif t == "remark":
            ws.merge_cells(start_row=n, start_column=2, end_row=n, end_column=4)
            ws.merge_cells(start_row=n, start_column=5, end_row=n, end_column=6)
            for j in range(1, 7): ws.cell(row=n, column=j).fill = REMARK_FILL
            ws.cell(row=n, column=1).font = Font(name="Calibri", size=10, italic=True)
            ws.cell(row=n, column=5).font = META
            ws.row_dimensions[n].height = 33
        elif t == "line" and not r["found"]:
            ws.cell(row=n, column=2).font = MUTEF
    safe = re.sub(r"\s+", "_", sub["clientName"])
    path = os.path.join(out_dir, f"{sub['reference']}_{safe}.xlsx")
    wb.save(path)
    return path

def upload_to_drive(drive, path, ref):
    meta = {"name": os.path.basename(path), "mimeType": "application/vnd.google-apps.spreadsheet"}
    if FOLDER_ID: meta["parents"] = [FOLDER_ID]
    media = MediaFileUpload(path, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    f = drive.files().create(body=meta, media_body=media, fields="id,webViewLink", supportsAllDrives=True).execute()
    return f.get("webViewLink", "")

# ---------------------------------------------------------------- CLI

def cmd_read_queue(a):
    sheets, _ = gauth()
    header, data, H = leads_header(sheets)
    out = []
    for i, vals in enumerate(data, start=2):
        g = lambda c: str(vals[H[c]]) if H.get(c) is not None and H[c] < len(vals) else ""
        status = g("Status")
        stale = False
        if status == ST["ENRICHING"]:
            try:
                age = (datetime.now(timezone.utc) - datetime.fromisoformat(g("Last Run").replace("Z","+00:00"))).total_seconds()/60
                stale = age > STALE_MINUTES
            except Exception:
                stale = True
        if status in (ST["QUEUED"], ST["CONFIRMED"]) or (status == ST["ENRICHING"] and stale):
            try: anchor_obj = json.loads(g("Identity Anchor") or "null")
            except Exception: anchor_obj = None
            confirmed = bool(anchor_obj and anchor_obj.get("confirmed"))
            phase = "research" if (status == ST["CONFIRMED"] or confirmed) else "identity"
            sub = dict(clientName=g("Client Name"), company=g("Company"), role=g("Role"),
                       industry=g("Industry"), officeCity=g("Based At"), website=g("Website"),
                       linkedin=g("LinkedIn"), budget=g("Budget"), area=g("Floor Area"),
                       purpose=g("Intended Use"), currentOffice=g("Office Today"),
                       leadSource=g("Lead Source"), existing=("Yes" if g("Lead Source") == "Existing Lodha Customer" else ""), notes=g("Notes"),
                       cpFirm=g("CP Firm"), cpSpoc=g("CP SPOC"), reapId=g("Reap ID"),
                       smName=g("SM Name"), smEmail=g("Owner Email"), reference=g("Reference"))
            out.append(dict(row=i, reference=sub["reference"], sub=sub, was_stale=stale,
                            phase=phase, anchor=anchor_obj if confirmed else None))
    print(json.dumps(dict(rows=out)))

def cmd_identity(a):
    """End of phase one: store the preliminary match and park the lead at
    IDENTITY_CHECK for the sourcing manager to confirm. Nothing else runs on
    this lead until they do."""
    sheets, _ = gauth()
    _, _, H = leads_header(sheets)
    anchor = json.loads(a.anchor) if a.anchor else dict(resolved=False, confidence=0, note="identity step returned nothing")
    anchor["checkedAt"] = now_iso()
    try:
        vals = read_all(sheets, LEADS)[a.row - 1]
        inter = json.loads(vals[H["Interactions"]]) if H.get("Interactions") is not None and H["Interactions"] < len(vals) and vals[H["Interactions"]] else []
        name = vals[H["Client Name"]] if H["Client Name"] < len(vals) else ""
    except Exception:
        inter, name = [], ""
    cfg = load_config(sheets)
    matched = bool(anchor.get("resolved")) and float(anchor.get("confidence") or 0) >= cfg["identityMin"]
    inter.append(dict(t=now_iso(), who="System", text=(
        f"Client identified: {anchor.get('full_name') or name}" + (f" \u2014 {anchor['company']}" if anchor.get("company") else "") + ". Awaiting confirmation."
        if matched else "Client not identified with confidence. Awaiting confirmation or amended details.")))
    a1 = lambda c: f"{LEADS}!{col(H[c])}{a.row}"
    write_cells(sheets, [
        (a1("Identity Anchor"), json.dumps(anchor)),
        (a1("Identity Confidence"), str(anchor.get("confidence")) if anchor.get("confidence") is not None else "not resolved"),
        (a1("Interactions"), json.dumps(inter)),
        (a1("Status"), ST["IDENTITY_CHECK"]), (a1("Last Run"), now_iso()),
    ])
    print(json.dumps(dict(ok=True, parked=True)))

def cmd_claim(a):
    sheets, _ = gauth()
    _, _, H = leads_header(sheets)
    write_cells(sheets, [(f"{LEADS}!{col(H['Status'])}{a.row}", ST["ENRICHING"]),
                        (f"{LEADS}!{col(H['Last Run'])}{a.row}", now_iso())])
    print(json.dumps(dict(ok=True)))

def cmd_evaluate(a):
    sheets, _ = gauth()
    cfg = load_config(sheets); cfg["anchors"] = load_anchors(sheets)
    sub = json.loads(a.sub); anchor = json.loads(a.anchor) if a.anchor else None
    ledger = json.loads(a.ledger)
    gates = evaluate_gates(ledger, anchor, sub.get("budget",""), cfg, sub)
    print(json.dumps(dict(gates=gates, verdict=rate(gates), breakdown=breakdown(gates, ledger, anchor))))

def cmd_workbook(a):
    sheets, drive = gauth()
    sub = json.loads(a.sub); anchor = json.loads(a.anchor) if a.anchor else {}
    ledger = json.loads(a.ledger); notes = json.loads(a.notes) if a.notes else {}
    path = write_workbook(build_rows(sub, anchor, ledger, notes), sub)
    try:
        url = upload_to_drive(drive, path, sub["reference"])
    except Exception as e:
        print(json.dumps(dict(url="", local_path=path, error=str(e)))); return
    print(json.dumps(dict(url=url)))

def cmd_writeback(a):
    sheets, _ = gauth()
    _, _, H = leads_header(sheets)
    gaps = json.loads(a.gaps) if a.gaps else []
    try:
        vals = read_all(sheets, LEADS)[a.row - 1]
        inter = json.loads(vals[H["Interactions"]]) if H.get("Interactions") is not None and H["Interactions"] < len(vals) and vals[H["Interactions"]] else []
    except Exception:
        inter = []
    inter.append(dict(t=now_iso(), who="System", text="Profile enriched and sent for approval."))
    a1 = lambda c: f"{LEADS}!{col(H[c])}{a.row}"
    write_cells(sheets, [
        (a1("Rating"), a.verdict), (a1("Identity Confidence"), a.confidence),
        (a1("Profile Summary"), a.summary),
        # the column keeps its name; it now carries the breakdown JSON
        (a1("Why This Rating"), a.breakdown if a.breakdown else (a.reasoning or "")),
        (a1("Meeting Should Establish"), "\n".join("• " + g for g in trim_gaps(gaps))),
        (a1("Escalations"), a.escalations or ""), (a1("Evidence Report"), a.evidence_url or ""),
        (a1("Interactions"), json.dumps(inter)), (a1("Status"), ST["PENDING"]),
        (a1("Last Run"), now_iso()),
    ])
    print(json.dumps(dict(ok=True)))

def cmd_fail(a):
    sheets, _ = gauth()
    _, _, H = leads_header(sheets)
    write_cells(sheets, [(f"{LEADS}!{col(H['Status'])}{a.row}", ST["FAILED"]),
                        (f"{LEADS}!{col(H['Run Log'])}{a.row}", ("FAILED: " + a.reason)[:400])])
    print(json.dumps(dict(ok=True)))

def main():
    p = argparse.ArgumentParser()
    sub = p.add_subparsers(dest="cmd", required=True)
    sub.add_parser("read_queue").set_defaults(fn=cmd_read_queue)
    c = sub.add_parser("claim"); c.add_argument("--row", type=int, required=True); c.set_defaults(fn=cmd_claim)
    c = sub.add_parser("identity"); c.add_argument("--row", type=int, required=True)
    c.add_argument("--anchor", default=""); c.set_defaults(fn=cmd_identity)
    c = sub.add_parser("evaluate")
    c.add_argument("--sub", required=True); c.add_argument("--anchor", default="")
    c.add_argument("--ledger", required=True); c.set_defaults(fn=cmd_evaluate)
    c = sub.add_parser("workbook")
    c.add_argument("--sub", required=True); c.add_argument("--anchor", default="")
    c.add_argument("--ledger", required=True); c.add_argument("--notes", default="")
    c.set_defaults(fn=cmd_workbook)
    c = sub.add_parser("writeback")
    c.add_argument("--row", type=int, required=True); c.add_argument("--verdict", required=True)
    c.add_argument("--confidence", default=""); c.add_argument("--summary", required=True)
    c.add_argument("--reasoning", default=""); c.add_argument("--breakdown", default="")
    c.add_argument("--gaps", default="[]")
    c.add_argument("--escalations", default=""); c.add_argument("--evidence_url", default="")
    c.set_defaults(fn=cmd_writeback)
    c = sub.add_parser("fail"); c.add_argument("--row", type=int, required=True)
    c.add_argument("--reason", required=True); c.set_defaults(fn=cmd_fail)
    if not SHEET_ID:
        print("GCR_SHEET_ID is not set", file=sys.stderr); sys.exit(2)
    a = p.parse_args(); a.fn(a)

if __name__ == "__main__":
    main()
